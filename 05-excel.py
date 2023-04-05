import openpyxl

wb = openpyxl.load_workbook('planilla.xlsx')

hoja = wb.active

wb.create_sheet("Sheet3")

print(wb.sheetnames)
